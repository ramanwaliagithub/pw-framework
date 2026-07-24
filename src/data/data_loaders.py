"""
Test Data Loaders.

Design pattern: Strategy (one loader per file format, uniform return shape
of list[dict] so parametrize/fixtures don't care which format backed the
data) + Facade (`load_test_data` picks the right loader by extension).

Why not just json.load() inline in every test? Because at 10k tests, data
sourcing needs to be swappable — today it's a checked-in JSON file, next
sprint it might be pulled from a shared Excel sheet the business team
maintains, or a REST endpoint. Tests call `load_test_data(path)` and never
change when the backing format does.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


def load_json(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_yaml(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_csv(path: str | Path) -> list[dict[str, str]]:
    with open(path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def load_excel(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """First row is treated as the header; each subsequent row becomes a dict."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    records = [dict(zip(headers, row)) for row in rows_iter]
    workbook.close()
    return records


_LOADERS = {
    ".json": load_json,
    ".yaml": load_yaml,
    ".yml": load_yaml,
    ".csv": load_csv,
    ".xlsx": load_excel,
}


def load_test_data(path: str | Path) -> Any:
    """Dispatch to the right loader based on file extension."""
    p = Path(path)
    loader = _LOADERS.get(p.suffix.lower())
    if not loader:
        raise ValueError(f"Unsupported test data format: {p.suffix}. Supported: {list(_LOADERS)}")
    return loader(p)
